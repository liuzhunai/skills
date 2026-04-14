#!/usr/bin/env python3
"""
58同城租房搜索 - 完整执行脚本
整合参数解析、URL构建、数据处理、Excel生成功能
"""

import re
import json
import math
import uuid
import os
import shutil
from datetime import datetime, timedelta
from typing import List, Dict, Optional, Tuple
from dataclasses import dataclass, asdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


@dataclass
class SearchParams:
    """搜索参数"""

    location: str
    radius_km: float = 5.0
    days: int = 15
    price_min: Optional[int] = None
    price_max: Optional[int] = None
    area_min: Optional[float] = None  # 最小面积
    area_max: Optional[float] = None  # 最大面积
    rooms: Optional[int] = None  # 1=一居, 2=两居, 3=三居
    rental_type: Optional[str] = None  # "整租" 或 "合租"
    city: str = "北京"
    city_code: str = "bj"


# 北京区域坐标
BEIJING_DISTRICTS = {
    "朝阳": {"url": "chaoyang", "lat": 39.9219, "lng": 116.4431},
    "海淀": {"url": "haidian", "lat": 39.9563, "lng": 116.3103},
    "昌平": {"url": "changping", "lat": 40.2206, "lng": 116.2311},
    "丰台": {"url": "fengtai", "lat": 39.8584, "lng": 116.2863},
    "大兴": {"url": "daxing", "lat": 39.7268, "lng": 116.3412},
    "通州": {"url": "tongzhouqu", "lat": 39.9066, "lng": 116.6627},
    "房山": {"url": "fangshan", "lat": 39.7488, "lng": 115.9938},
    "顺义": {"url": "shunyi", "lat": 40.1301, "lng": 116.6545},
    "西城": {"url": "xicheng", "lat": 39.9128, "lng": 116.3662},
    "东城": {"url": "dongcheng", "lat": 39.9289, "lng": 116.4162},
    "石景山": {"url": "shijingshan", "lat": 39.9063, "lng": 116.1954},
}

# 常见地点坐标
LOCATION_COORDS = {
    "百度科技园": {"lat": 40.0569, "lng": 116.3015, "district": "海淀"},
    "百度大厦": {"lat": 40.0569, "lng": 116.3015, "district": "海淀"},
    "中关村": {"lat": 39.9841, "lng": 116.3075, "district": "海淀"},
    "国贸": {"lat": 39.9087, "lng": 116.4594, "district": "朝阳"},
    "望京": {"lat": 40.0032, "lng": 116.4773, "district": "朝阳"},
    "西二旗": {"lat": 40.0569, "lng": 116.3015, "district": "海淀"},
    "上地": {"lat": 40.0310, "lng": 116.2870, "district": "海淀"},
    "五道口": {"lat": 39.9929, "lng": 116.3370, "district": "海淀"},
    "三里屯": {"lat": 39.9324, "lng": 116.4536, "district": "朝阳"},
    "王府井": {"lat": 39.9139, "lng": 116.4103, "district": "东城"},
    "西单": {"lat": 39.9134, "lng": 116.3728, "district": "西城"},
    "北京南站": {"lat": 39.8654, "lng": 116.3786, "district": "丰台"},
    "北京西站": {"lat": 39.8946, "lng": 116.3220, "district": "丰台"},
}

# 海淀区各地点坐标（用于距离估算）
HAIDIAN_LOCATIONS = {
    "公主坟": (39.9006, 116.3211),
    "马连洼": (40.0289, 116.2822),
    "小西天": (39.9489, 116.3644),
    "花园桥": (39.9378, 116.3433),
    "海淀周边": (40.0300, 116.2500),
    "苏家坨": (40.0889, 116.1511),
    "田村": (39.9389, 116.2211),
    "上庄": (40.1111, 116.1300),
    "永定路": (39.9078, 116.2433),
    "皂君庙": (39.9533, 116.3356),
    "西二旗": (40.0569, 116.3015),
    "北洼路": (39.9311, 116.3033),
    "军博": (39.9089, 116.3256),
    "四季青": (39.9589, 116.2433),
    "清河": (40.0333, 116.3489),
    "西北旺": (40.0444, 116.2733),
    "西三旗": (40.0667, 116.3333),
    "五道口": (39.9929, 116.3370),
    "中关村": (39.9841, 116.3075),
    "上地": (40.0310, 116.2870),
    "安宁庄": (40.0550, 116.3189),
    "回龙观": (40.0711, 116.3150),
    "龙泽": (40.0633, 116.3089),
}

# 北京地铁站坐标（部分常用站点）
BEIJING_SUBWAY_STATIONS = {
    # 13号线
    "西二旗": (40.0569, 116.3015),
    "上地": (40.0310, 116.2870),
    "五道口": (39.9929, 116.3370),
    "知春路": (39.9756, 116.3394),
    "芍药居": (39.9672, 116.4283),
    "望京西": (40.0022, 116.4194),
    "东直门": (39.9444, 116.4344),
    # 昌平线
    "昌平": (40.2161, 116.2311),
    "沙河": (40.1489, 116.2911),
    "生命科学园": (40.0889, 116.2911),
    # 16号线
    "西北旺": (40.0444, 116.2733),
    "马连洼": (40.0289, 116.2822),
    "西苑": (40.0061, 116.2839),
    # 4号线
    "中关村": (39.9841, 116.3075),
    "海淀黄庄": (39.9806, 116.3125),
    "人民大学": (39.9656, 116.3228),
    "魏公村": (39.9533, 116.3267),
    "国家图书馆": (39.9333, 116.3194),
    # 10号线
    "知春路": (39.9756, 116.3394),
    "海淀黄庄": (39.9806, 116.3125),
    "苏州街": (39.9711, 116.3061),
    "巴沟": (39.9644, 116.2894),
    "慈寿寺": (39.9489, 116.2794),
    # 1号线
    "五棵松": (39.9089, 116.2567),
    "万寿路": (39.9061, 116.2739),
    "公主坟": (39.9006, 116.3211),
    "军事博物馆": (39.9089, 116.3256),
    # 6号线
    "海淀五路居": (39.9267, 116.2639),
    "慈寿寺": (39.9489, 116.2794),
    "花园桥": (39.9378, 116.3433),
    # 8号线
    "永泰庄": (40.0239, 116.3489),
    "西小口": (40.0361, 116.3511),
    "育新": (40.0489, 116.3511),
    # 其他
    "回龙观": (40.0711, 116.3150),
    "龙泽": (40.0633, 116.3089),
    "霍营": (40.0711, 116.3711),
}


def parse_query(query: str) -> SearchParams:
    """解析用户输入的查询"""
    params = SearchParams(location="")

    # 提取距离
    radius_match = re.search(r"(\d+(?:\.\d+)?)\s*(km|公里|米|m)", query)
    if radius_match:
        value = float(radius_match.group(1))
        unit = radius_match.group(2).lower()
        params.radius_km = value if unit in ["km", "公里"] else value / 1000

    # 提取时间
    days_match = re.search(r"(\d+)\s*天", query)
    if days_match:
        params.days = int(days_match.group(1))

    # 提取价格区间
    price_match = re.search(r"价格?(\d+)-(\d+)", query)
    if price_match:
        params.price_min = int(price_match.group(1))
        params.price_max = int(price_match.group(2))
    else:
        price_max_match = re.search(r"(\d+)以[内下]", query)
        if price_max_match:
            params.price_max = int(price_max_match.group(1))

    # 提取面积区间（支持：面积30-50、30-50平、50平以内、50平米以下）
    area_match = re.search(
        r"面积?(\d+(?:\.\d+)?)-(\d+(?:\.\d+)?)\s*(?:平|平米)?", query
    )
    if area_match:
        params.area_min = float(area_match.group(1))
        params.area_max = float(area_match.group(2))
    else:
        area_max_match = re.search(r"(\d+(?:\.\d+)?)\s*(?:平|平米)\s*以[内下]", query)
        if area_max_match:
            params.area_max = float(area_max_match.group(1))

    # 提取户型
    if "一居" in query or "一室" in query or "开间" in query:
        params.rooms = 1
    elif "两居" in query or "两室" in query:
        params.rooms = 2
    elif "三居" in query or "三室" in query:
        params.rooms = 3

    # 提取租赁方式
    if "整租" in query:
        params.rental_type = "整租"
    elif "合租" in query:
        params.rental_type = "合租"

    # 提取地点关键词
    location_query = query
    location_query = re.sub(r"(\d+(?:\.\d+)?)\s*(km|公里|米|m)", "", location_query)
    location_query = re.sub(r"(\d+)\s*天[内以]?", "", location_query)
    location_query = re.sub(r"价格?(\d+)-(\d+)", "", location_query)
    location_query = re.sub(r"(\d+)以[内下]", "", location_query)
    location_query = re.sub(
        r"面积?(\d+(?:\.\d+)?)-(\d+(?:\.\d+)?)\s*(?:平|平米)?", "", location_query
    )
    location_query = re.sub(
        r"(\d+(?:\.\d+)?)\s*(?:平|平米)\s*以[内下]", "", location_query
    )
    location_query = re.sub(
        r"(一居|两居|三居|一室|两室|三室|开间|整租|合租)", "", location_query
    )
    location_query = re.sub(r"(附近|周边|的|房源|租房|房子|个人)", "", location_query)
    params.location = location_query.strip()

    return params


def get_location_coords(location: str) -> Optional[Dict]:
    """获取地点坐标"""
    if location in LOCATION_COORDS:
        return LOCATION_COORDS[location]
    for name, coords in LOCATION_COORDS.items():
        if location in name or name in location:
            return coords
    for district, info in BEIJING_DISTRICTS.items():
        if district in location:
            return {"lat": info["lat"], "lng": info["lng"], "district": district}
    return None


def calculate_distance(lat1: float, lng1: float, lat2: float, lng2: float) -> float:
    """计算两点之间的距离（km）"""
    R = 6371
    lat1_rad, lat2_rad = math.radians(lat1), math.radians(lat2)
    delta_lat = math.radians(lat2 - lat1)
    delta_lng = math.radians(lng2 - lng1)
    a = (
        math.sin(delta_lat / 2) ** 2
        + math.cos(lat1_rad) * math.cos(lat2_rad) * math.sin(delta_lng / 2) ** 2
    )
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def get_nearby_districts(lat: float, lng: float, radius_km: float) -> List[Dict]:
    """获取附近的行政区划"""
    nearby = []
    for name, info in BEIJING_DISTRICTS.items():
        dist = calculate_distance(lat, lng, info["lat"], info["lng"])
        if dist <= radius_km + 15:
            nearby.append({"name": name, "url": info["url"], "distance": dist})
    return sorted(nearby, key=lambda x: x["distance"])


def build_search_url(
    city_code: str,
    district_url: str,
    price_min: int = None,
    price_max: int = None,
    rooms: int = None,
    rental_type: str = None,
    page: int = 1,
) -> str:
    """构建58同城搜索URL（个人房源）

    Args:
        city_code: 城市代码 (如 bj=北京)
        district_url: 区域URL名 (如 haidian=海淀)
        price_min: 最低价格
        price_max: 最高价格
        rooms: 户型 (1=一居, 2=两居, 3=三居)
        rental_type: 租赁方式 ("整租" 或 "合租")
        page: 页码

    Returns:
        58同城搜索URL
    """
    # 根据租赁方式选择URL路径
    if rental_type == "合租":
        path = "hezu"
    else:
        path = "zufang"

    url = f"https://{city_code}.58.com/{district_url}/{path}/0/"

    if rooms:
        url = url.rstrip("/") + f"/j{rooms}/"

    price_code = None
    if price_max:
        if price_max <= 1000:
            price_code = 9
        elif price_max <= 1500:
            price_code = 10
        elif price_max <= 2000:
            price_code = 11
        elif price_max <= 3000:
            price_code = 12
        elif price_max <= 5000:
            price_code = 13
        elif price_max <= 8000:
            price_code = 14
        else:
            price_code = 15

    if price_code:
        url = url.rstrip("/") + f"/b{price_code}/"

    if page > 1:
        url = url.rstrip("/") + f"/pn{page}/"

    return url


def estimate_location_coords(location_text: str) -> Optional[Tuple[float, float]]:
    """根据位置文本估算坐标"""
    for loc_name, coords in HAIDIAN_LOCATIONS.items():
        if loc_name in location_text:
            return coords
    return None


def find_nearest_subway(
    lat: float, lng: float, max_distance_km: float = 3.0
) -> Optional[str]:
    """查找最近的地铁站"""
    nearest_station = None
    min_distance = float("inf")

    for station_name, coords in BEIJING_SUBWAY_STATIONS.items():
        distance = calculate_distance(lat, lng, coords[0], coords[1])
        if distance < min_distance:
            min_distance = distance
            nearest_station = station_name

    if min_distance <= max_distance_km:
        return f"{nearest_station} ({int(min_distance * 1000)}m)"
    return None


def filter_listings_by_distance(
    raw_listings: List[Dict],
    target_lat: float,
    target_lng: float,
    radius_km: float,
    area_min: float = None,
    area_max: float = None,
) -> List[Dict]:
    """按距离和面积过滤房源

    Args:
        raw_listings: 原始房源列表
        target_lat: 目标纬度
        target_lng: 目标经度
        radius_km: 搜索半径(km)
        area_min: 最小面积(㎡)
        area_max: 最大面积(㎡)

    Returns:
        过滤后的房源列表，按距离排序
    """
    filtered = []
    for listing in raw_listings:
        location_text = listing.get("location", "")
        coords = estimate_location_coords(location_text)
        if coords:
            distance = calculate_distance(target_lat, target_lng, coords[0], coords[1])
            if distance <= radius_km:
                # 面积过滤
                area_str = listing.get("area", "")
                area_match = re.search(r"(\d+(?:\.\d+)?)", str(area_str))
                if area_match:
                    area = float(area_match.group(1))
                    if area_min and area < area_min:
                        continue
                    if area_max and area > area_max:
                        continue

                listing["distance"] = round(distance, 2)
                # 查找最近地铁站
                nearest_subway = find_nearest_subway(coords[0], coords[1])
                listing["nearest_subway"] = nearest_subway or ""
                filtered.append(listing)
    return sorted(filtered, key=lambda x: x["distance"])


def generate_excel(listings: List[Dict], output_path: str, location_name: str) -> str:
    """生成Excel文件"""
    tmp_dir = f".dumate/xlsx-{uuid.uuid4()}"
    os.makedirs(tmp_dir, exist_ok=True)

    wb = Workbook()
    sheet = wb.active
    sheet.title = "房源列表"

    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(bold=True, size=11, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = [
        "标题",
        "类型",
        "小区",
        "房间数",
        "面积(㎡)",
        "租金(元/月)",
        "距离(km)",
        "最近地铁站",
        "发布时间",
        "链接",
    ]
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for row_idx, listing in enumerate(listings, 2):
        values = [
            listing.get("title", ""),
            listing.get("rental_type", "整租"),
            listing.get("community", ""),
            listing.get("rooms", 1),
            listing.get("area", ""),
            listing.get("price", ""),
            listing.get("distance", ""),
            listing.get("nearest_subway", ""),
            listing.get("publish_time", ""),
            listing.get("url", ""),
        ]
        for col_idx, value in enumerate(values, 1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = center_align if col_idx in [4, 5, 6, 7] else left_align

    column_widths = {
        "A": 35,
        "B": 8,
        "C": 20,
        "D": 8,
        "E": 10,
        "F": 12,
        "G": 10,
        "H": 18,
        "I": 12,
        "J": 50,
    }
    for col_letter, width in column_widths.items():
        sheet.column_dimensions[col_letter].width = width

    tmp_file = os.path.join(tmp_dir, f"{location_name}附近房源.xlsx")
    wb.save(tmp_file)

    final_path = output_path or f"{location_name}附近房源.xlsx"
    shutil.copy2(tmp_file, final_path)
    shutil.rmtree(tmp_dir, ignore_errors=True)

    return final_path


def main():
    import argparse

    parser = argparse.ArgumentParser(description="58同城租房搜索")
    parser.add_argument(
        "query", nargs="?", default="中关村附近5km的房源", help="搜索关键词"
    )
    parser.add_argument("--output", default=None, help="输出Excel文件路径")
    parser.add_argument("--json", action="store_true", help="输出JSON格式")
    args = parser.parse_args()

    params = parse_query(args.query)
    print(f"解析结果: {asdict(params)}")

    location_info = get_location_coords(params.location)
    if not location_info:
        print(f"警告: 未找到地点 '{params.location}'，使用默认坐标")
        location_info = {"lat": 39.9042, "lng": 116.4074, "district": "东城"}

    print(
        f"目标地点: {params.location} ({location_info['lat']}, {location_info['lng']})"
    )

    districts = get_nearby_districts(
        location_info["lat"], location_info["lng"], params.radius_km
    )
    print(f"相关区域: {[d['name'] for d in districts[:5]]}")

    search_urls = []
    for district in districts[:3]:
        url = build_search_url(
            params.city_code,
            district["url"],
            params.price_min,
            params.price_max,
            params.rooms,
            params.rental_type,
        )
        search_urls.append(
            {"district": district["name"], "url": url, "distance": district["distance"]}
        )

    print(f"搜索URL: {search_urls[0]['url']}")

    config = {
        "params": asdict(params),
        "location": location_info,
        "districts": districts,
        "search_urls": search_urls,
        "created_at": datetime.now().isoformat(),
    }

    if args.json:
        print(json.dumps(config, ensure_ascii=False, indent=2))
    else:
        with open("search_config.json", "w", encoding="utf-8") as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        print(f"\n配置已保存到: search_config.json")

    print(f"\n下一步: 使用 dumate-browser-use 访问搜索URL并提取房源数据")


if __name__ == "__main__":
    main()
