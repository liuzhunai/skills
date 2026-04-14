"""Excel exporter implementation."""

import os
import uuid
import shutil
from typing import List, Dict, Optional

from .base import Exporter
from ..models.listing import Listing


class ExcelExporter(Exporter):
    """Excel导出器

    使用openpyxl生成格式化的Excel文件。
    """

    # 默认列宽配置
    COLUMN_WIDTHS = {
        "A": 35,  # 标题
        "B": 8,  # 类型
        "C": 20,  # 小区
        "D": 8,  # 房间数
        "E": 10,  # 面积
        "F": 12,  # 租金
        "G": 10,  # 距离
        "H": 18,  # 最近地铁站
        "I": 12,  # 发布时间
        "J": 50,  # 链接
    }

    # 表头
    HEADERS = [
        "标题",
        "类型",
        "小区",
        "房间数",
        "面积(㎡)",
        "租金(元/月)",
        "距离(km)",
        "最近地铁站",
        "发布时间",
        "链接",
    ]

    def __init__(self):
        self._workbook = None
        self._sheet = None

    def export(
        self,
        listings: List[Listing],
        output_path: str,
        location_name: str = "房源",
        **kwargs,
    ) -> str:
        """导出房源数据到Excel

        Args:
            listings: 房源列表
            output_path: 输出路径
            location_name: 地点名称（用于文件名）

        Returns:
            实际输出路径
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        # 创建临时目录
        tmp_dir = f".dumate/xlsx-{uuid.uuid4()}"
        os.makedirs(tmp_dir, exist_ok=True)

        # 创建工作簿
        wb = Workbook()
        sheet = wb.active
        sheet.title = "房源列表"

        # 样式定义
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(bold=True, size=11, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # 写入表头
        for col, header in enumerate(self.HEADERS, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # 写入数据
        for row_idx, listing in enumerate(listings, 2):
            values = self._listing_to_row(listing)
            for col_idx, value in enumerate(values, 1):
                cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = center_align if col_idx in [4, 5, 6, 7] else left_align

        # 设置列宽
        for col_letter, width in self.COLUMN_WIDTHS.items():
            sheet.column_dimensions[col_letter].width = width

        # 保存临时文件
        tmp_file = os.path.join(tmp_dir, f"{location_name}附近房源.xlsx")
        wb.save(tmp_file)

        # 复制到目标路径
        final_path = output_path or f"{location_name}附近房源.xlsx"
        shutil.copy2(tmp_file, final_path)

        # 清理临时目录
        shutil.rmtree(tmp_dir, ignore_errors=True)

        return final_path

    def _listing_to_row(self, listing: Listing) -> List:
        """将房源转换为行数据"""
        return [
            listing.title,
            listing.rental_type,
            listing.community,
            listing.rooms,
            listing.area,
            listing.price,
            listing.distance,
            listing.nearest_subway or "",
            listing.publish_time or "",
            listing.url,
        ]

    def get_format(self) -> str:
        return "xlsx"

    @classmethod
    def create_with_custom_headers(
        cls, headers: List[str], column_widths: Dict[str, int] = None
    ) -> "ExcelExporter":
        """创建自定义表头的导出器

        Args:
            headers: 自定义表头
            column_widths: 自定义列宽

        Returns:
            配置好的导出器实例
        """
        exporter = cls()
        if headers:
            exporter.HEADERS = headers
        if column_widths:
            exporter.COLUMN_WIDTHS = column_widths
        return exporter
